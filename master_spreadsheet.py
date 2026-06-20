"""
master_spreadsheet.py - 将已分类账单追加到年度总表 xlsx
"""

from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd

MAIN_COLUMNS = [
    'Name', 'Category', 'Amount', 'Date', 'Person', 'Source', '是否自动分类',
]

MONTH_MAP = {
    1: '1月', 2: '2月', 3: '3月', 4: '4月', 5: '5月', 6: '6月',
    7: '7月', 8: '8月', 9: '9月', 10: '10月', 11: '11月', 12: '12月',
}


@dataclass
class MergeResult:
    success: bool
    merged_count: int = 0
    skipped_count: int = 0
    sheet_name: str = ''
    path: str = ''
    error: str = ''

    def summary(self) -> str:
        if not self.success:
            return f'总表合并失败: {self.error}'
        return (
            f'总表 {self.sheet_name}：新增 {self.merged_count} 条，'
            f'跳过重复 {self.skipped_count} 条'
        )


class MasterSpreadsheetMerger:
    """将导出 DataFrame 按月份 sheet 追加到年度总表，支持 Date+Amount+Name 去重。"""

    def __init__(self, config_manager):
        self.config = config_manager

    def is_enabled(self) -> bool:
        return bool(self.config.get('master_spreadsheet.enabled', False))

    def merge_if_enabled(self, df: pd.DataFrame, force: bool = False) -> Optional[MergeResult]:
        if not force and not self.is_enabled():
            return None
        return self.merge_dataframe(df)

    def merge_dataframe(self, df: pd.DataFrame) -> MergeResult:
        if df is None or len(df) == 0:
            return MergeResult(success=False, error='没有可合并的数据')

        cfg = self._get_settings()
        dedupe_keys = cfg.get('dedupe_keys') or ['Date', 'Amount', 'Name']
        missing = [col for col in dedupe_keys if col not in df.columns]
        if missing:
            return MergeResult(success=False, error=f'缺少去重字段: {", ".join(missing)}')

        month_label = extract_bill_month_label(df)
        year_label = extract_bill_year(df)
        path = self._resolve_path(cfg.get('path', '已分类/{year}/{year}总表.xlsx'), year_label)
        sheet_name = (cfg.get('sheet_naming', '{month}') or '{month}').format(
            month=month_label, year=year_label
        )

        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            return MergeResult(success=False, error='缺少 openpyxl 依赖')

        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            if path.exists() and cfg.get('backup_before_merge', True):
                backup = path.with_suffix(path.suffix + '.bak')
                try:
                    shutil.copy2(path, backup)
                except OSError:
                    pass

            if path.exists():
                wb = load_workbook(path)
            else:
                wb = Workbook()
                if wb.sheetnames == ['Sheet'] and wb.active.title == 'Sheet':
                    wb.remove(wb.active)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append(MAIN_COLUMNS)

            header_map = self._build_header_map(ws)
            if not header_map:
                for col_idx, name in enumerate(MAIN_COLUMNS, start=1):
                    ws.cell(row=1, column=col_idx, value=name)
                header_map = self._build_header_map(ws)

            existing_keys = self._read_existing_keys(ws, header_map, dedupe_keys)
            merged_count = 0
            skipped_count = 0

            for _, row in df.iterrows():
                key = self._make_dedupe_key(row, dedupe_keys)
                if key in existing_keys:
                    skipped_count += 1
                    continue
                existing_keys.add(key)
                ws.append(self._row_values(row, header_map))
                merged_count += 1

            wb.save(path)
            print(f'✅ {MergeResult(True, merged_count, skipped_count, sheet_name, str(path)).summary()}')
            return MergeResult(
                success=True,
                merged_count=merged_count,
                skipped_count=skipped_count,
                sheet_name=sheet_name,
                path=str(path),
            )
        except Exception as exc:
            return MergeResult(success=False, error=format_merge_error(exc, path))

    def _get_settings(self) -> Dict[str, Any]:
        return self.config.get('master_spreadsheet', {}) or {}

    def _resolve_path(self, rel_path: str, year: str = '') -> Path:
        text = (rel_path or '').format(year=year or '')
        path = Path(text)
        if path.is_absolute():
            return path
        try:
            from app_paths import get_app_dir
            return get_app_dir() / path
        except ImportError:
            return Path(self.config.config_dir) / path

    def _sheet_name_for_df(self, df: pd.DataFrame, pattern: str) -> str:
        month_label = extract_bill_month_label(df)
        year = extract_bill_year(df)
        return pattern.format(month=month_label, year=year)

    def resolve_export_dir(self, df: pd.DataFrame) -> Path:
        template = self.config.get('files.export_dir', '已分类/{year}')
        year = extract_bill_year(df)
        rel = (template or '已分类/{year}').format(year=year)
        return self._resolve_path(rel, year)

    def _build_header_map(self, ws) -> Dict[str, int]:
        header_map: Dict[str, int] = {}
        if ws.max_row < 1:
            return header_map
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col_idx).value
            if value is None:
                continue
            header_map[str(value).strip()] = col_idx
        return header_map

    def _read_existing_keys(
        self,
        ws,
        header_map: Dict[str, int],
        dedupe_keys: List[str],
    ) -> Set[Tuple[Any, ...]]:
        keys: Set[Tuple[Any, ...]] = set()
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for key in dedupe_keys:
                col_idx = header_map.get(key)
                if not col_idx:
                    continue
                row_data[key] = ws.cell(row=row_idx, column=col_idx).value
            if row_data:
                keys.add(self._make_dedupe_key(row_data, dedupe_keys))
        return keys

    def _make_dedupe_key(self, row, dedupe_keys: List[str]) -> Tuple[Any, ...]:
        parts = []
        for key in dedupe_keys:
            value = row.get(key) if isinstance(row, dict) else row[key]
            if key == 'Amount':
                parts.append(normalize_amount(value))
            elif key == 'Date':
                parts.append(normalize_date(value))
            else:
                parts.append(normalize_text(value))
        return tuple(parts)

    def _row_values(self, row, header_map: Dict[str, int]) -> List[Any]:
        if not header_map:
            return [row.get(col) for col in MAIN_COLUMNS]
        max_col = max(header_map.values())
        values: List[Any] = [None] * max_col
        for col_name in MAIN_COLUMNS:
            col_idx = header_map.get(col_name)
            if col_idx:
                values[col_idx - 1] = row.get(col_name)
        return values


def normalize_amount(value) -> str:
    try:
        return f'{float(value):.2f}'
    except (TypeError, ValueError):
        return normalize_text(value)


def normalize_date(value) -> str:
    text = normalize_text(value)
    if not text:
        return text
    parsed = pd.to_datetime(text, errors='coerce')
    if pd.isna(parsed):
        return text
    return parsed.strftime('%Y-%m-%d')


def normalize_text(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ''
    return str(value).strip()


def format_merge_error(exc: Exception, path: Path) -> str:
    """将底层 IO 异常转为用户可理解的提示。"""
    errno = getattr(exc, 'errno', None)
    if isinstance(exc, PermissionError) or errno == 13:
        return (
            f'无法写入总表（文件可能正被 Excel 打开）：\n{path}\n\n'
            '请先关闭 Excel 中的该文件，再点击「同步本单到总表」重试。'
        )
    return str(exc)


def extract_bill_month_label(df: pd.DataFrame) -> str:
    """从 Date 列取众数月份，避免排序后首行导致月份错误。"""
    if 'Date' not in df.columns or len(df) == 0:
        from datetime import datetime
        return datetime.now().strftime('%m').lstrip('0') + '月'

    dates = pd.to_datetime(df['Date'], errors='coerce').dropna()
    if dates.empty:
        from datetime import datetime
        return datetime.now().strftime('%m').lstrip('0') + '月'

    month_counts = dates.dt.month.value_counts()
    month_num = int(month_counts.index[0])
    return MONTH_MAP.get(month_num, f'{month_num}月')


def extract_bill_year(df: pd.DataFrame) -> str:
    if 'Date' not in df.columns or len(df) == 0:
        from datetime import datetime
        return str(datetime.now().year)

    dates = pd.to_datetime(df['Date'], errors='coerce').dropna()
    if dates.empty:
        from datetime import datetime
        return str(datetime.now().year)

    year_counts = dates.dt.year.value_counts()
    return str(int(year_counts.index[0]))
