"""master_spreadsheet.py 单元测试"""

import pandas as pd
import pytest
from openpyxl import load_workbook

from config import ConfigManager
from data_loader import DataLoader
from master_spreadsheet import (
    MasterSpreadsheetMerger,
    extract_bill_month_label,
    format_merge_error,
    normalize_amount,
)


@pytest.fixture
def sample_df():
    return pd.DataFrame([
        {
            'Name': '商户A - 商品1',
            'Category': '餐饮',
            'Amount': -12.5,
            'Date': '2026-04-15',
            'Person': '袁程波',
            'Source': '微信',
            '是否自动分类': '是',
        },
        {
            'Name': '商户B',
            'Category': '出行',
            'Amount': -3.0,
            'Date': '2026-04-01',
            'Person': '袁程波',
            'Source': '支付宝',
            '是否自动分类': '否',
        },
    ])


def test_extract_bill_month_label_uses_mode(sample_df):
    assert extract_bill_month_label(sample_df) == '4月'


def test_normalize_amount():
    assert normalize_amount(-12.5) == '-12.50'
    assert normalize_amount('3') == '3.00'


def test_format_merge_error_permission_denied(tmp_path):
    msg = format_merge_error(PermissionError(13, 'denied'), tmp_path / '2026总表.xlsx')
    assert 'Excel' in msg
    assert '2026总表.xlsx' in msg


def test_merge_creates_sheet_and_appends(tmp_path, sample_df):
    config = ConfigManager(config_dir=str(tmp_path))
    config.set('master_spreadsheet.enabled', True)
    config.set('master_spreadsheet.path', str(tmp_path / '已分类/2026/2026总表.xlsx'))
    config.set('master_spreadsheet.backup_before_merge', False)

    merger = MasterSpreadsheetMerger(config)
    result = merger.merge_dataframe(sample_df)

    assert result.success is True
    assert result.merged_count == 2
    assert result.skipped_count == 0
    assert result.sheet_name == '4月'

    wb = load_workbook(result.path)
    ws = wb['4月']
    assert ws.max_row == 3


def test_merge_deduplicates_existing_rows(tmp_path, sample_df):
    config = ConfigManager(config_dir=str(tmp_path))
    config.set('master_spreadsheet.path', str(tmp_path / '2026总表.xlsx'))
    config.set('master_spreadsheet.backup_before_merge', False)

    merger = MasterSpreadsheetMerger(config)
    first = merger.merge_dataframe(sample_df)
    second = merger.merge_dataframe(sample_df)

    assert first.merged_count == 2
    assert second.success is True
    assert second.merged_count == 0
    assert second.skipped_count == 2


def test_detect_bill_source_from_path():
    assert DataLoader.detect_bill_source_from_path('原始账单/袁程波-支付宝-4月.csv') == '支付宝'
    assert DataLoader.detect_bill_source_from_path('原始账单/袁程波-微信-4月.xlsx') == '微信'
    assert DataLoader.detect_bill_source_from_path('other.csv') is None


def test_resolve_bill_source_prefers_filename(tmp_path):
    config = ConfigManager(config_dir=str(tmp_path))
    loader = DataLoader(config)
    resolved = loader.resolve_bill_source(
        str(tmp_path / '袁程波-支付宝交易明细.csv'),
        '微信',
    )
    assert resolved == '支付宝'
